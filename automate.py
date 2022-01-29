import openpyxl
import shutil
import datetime
from win32com import client
import os

in_wb = openpyxl.load_workbook('input/dnw.xlsx')
in_sheet = in_wb['dnw']

max_row = in_sheet.max_row
max_col = in_sheet.max_column

print(f"Max Row: {max_row}, Max Column: {max_col}")

row_no = 2

for i in range(1, max_row):
    # Reading data from the excel file
    account_number  = in_sheet['A' + str(i+1)].value
    phone_number = in_sheet['B' + str(i+1)].value
    owner_name = in_sheet['C' + str(i+1)].value
    address = in_sheet['D' + str(i+1)].value
    district = in_sheet['E' + str(i+1)].value
    sub_district = in_sheet['F' + str(i+1)].value
    xero_id = in_sheet['G' + str(i+1)].value
    product_name = in_sheet['H' + str(i+1)].value
    #print(f"Account Number: {account_number}\nPhone Number: {phone_number}\nOwner Name: {owner_name}\nAddress: {address}\nDistrict: {district}\nSub District: {sub_district}\nXero ID: {xero_id}\nProduct Name: {product_name}\n")

    # Creating output files from the template
    new_file_name = "output/" + str(account_number) + ".xlsx"
    shutil.copyfile('invoice_template.xlsx', new_file_name)

    # Updating output file's data
    invoice_id = "INV_" + str(xero_id)
    new_wb = openpyxl.load_workbook(new_file_name)
    new_sheet=new_wb.active
    new_sheet.title = invoice_id
    new_sheet['C1'] = str(invoice_id)
    new_sheet['I1'] = datetime.date.today().strftime("%d-%b-%Y")
    new_sheet['C2'] = str(owner_name)
    new_sheet['C3'] = str(phone_number)
    new_sheet['C4'] = str(address)
    new_sheet['C6'] = str(sub_district)
    new_sheet['C7'] = str(district)
    new_sheet['A12'] = str(account_number)
    new_sheet['B15'] = str(product_name)
    new_wb.save(new_file_name)

    # Converting to pdf
    in_excel_file = os.getcwd() + f"\\output\\{account_number}.xlsx"
    out_pdf_file = os.getcwd() + f"\\output\\{account_number}.pdf"
    #print(in_excel_file)
    excel = client.Dispatch("Excel.Application")
    #pdf_name = "output/" + str(account_number) + ".pdf"
    sheets = excel.Workbooks.Open(in_excel_file)
    work_sheets = sheets.Worksheets[0]
    work_sheets.ExportAsFixedFormat(0, out_pdf_file)
    excel.Application.Quit()
    print(f"File Created: {out_pdf_file}")


    

